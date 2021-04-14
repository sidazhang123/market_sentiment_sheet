from os import path
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Color, PatternFill


def _bold(data, ws):
    for col, c in enumerate(data):
        c = ws.cell(column=col + 1, row=1, value=c)
        c.font = Font(bold=True)
        yield c


# 4000只股中，多于2300只上涨/下跌，则认为是红绿盘, >3连板如有则标蓝
def _red_green_blue(data, ws, append_at, significance_threshold=2300):
    red = PatternFill(patternType='solid', fgColor=Color(rgb='00FF0000'))
    green = PatternFill(patternType='solid', fgColor=Color(rgb='0000FF00'))
    blue = PatternFill(patternType='solid', fgColor=Color(rgb='000000FF'))
    fill = dict()
    if data[1] > significance_threshold:
        fill[1]=red
    elif data[2] > significance_threshold:
        fill[2]=green
    if type(data[-2]) is int and data[-2]>0:
        fill[len(data)-2]=blue
    for col, c in enumerate(data):
        if col == 0:
            c = ws.cell(column=col + 1, row=append_at, value=c)
            c.font = Font(bold=True)
        if col in fill:
            c = ws.cell(column=col + 1, row=append_at, value=c)
            c.fill = fill[col]
        yield c

def _grey(data, ws, append_at):
    grey = PatternFill(patternType='solid', fgColor=Color(rgb='00969696'))
    for col, c in enumerate(data):
        if col == 0:
            c = ws.cell(column=col + 1, row=append_at, value=c)
            c.font = Font(bold=True)
        else:
            c = ws.cell(column=col + 1, row=append_at, value=c)
            c.fill = grey
        yield c

def find_gte_3(limit_record)->tuple:
    s=[]
    c=0
    for num in limit_record:
        if num>3:
            c+=len(limit_record[num])
            s.append('; '.join(j for j in ['-{}-'.format(num).join(i) for i in limit_record[num]]))
    return c,'; '.join(s)

# 日期   红盘 绿盘 最终涨停 最终跌停 炸板 1连板 2连板 3连板 个股（3连，股名+行业） 3连板以上 个股（股名+连板高度+行业）
# today,u_num, d_num, boom_num, limit_d_num, limit_u_num, limit_record
def write(today, info=None):

    today=today[:4]+"-"+today[4:6]+"-"+today[6:]
    fn = '市场情绪监控表.xlsx'
    header = ['日期', '红盘', '绿盘', '涨停', '跌停', '炸板', '1连', '2连', '3连', '个股', '>3连', '个股']
    if not path.exists(fn):
        wb = Workbook()
        ws = wb.active
        ws.title = "监控表"
        ws.append(_bold(header, ws))
        append_at = 2
    else:
        wb = load_workbook(fn)
        ws = wb.active
        append_at = len(list(ws.rows)) + 1
    # 非交易日
    if info is None:
        data=['']*len(header)
        data[0]=today
        ws.append(_grey(data, ws, append_at))
    else:
        u_num, d_num, boom_num, limit_d_num, limit_u_num, limit_record = info
        data = [today, u_num, d_num, limit_u_num, limit_d_num, boom_num,
                len(limit_record[1]) if 1 in limit_record else '',
                len(limit_record[2]) if 2 in limit_record else '',
                len(limit_record[3]) if 3 in limit_record else '',
                '; '.join(j for j in ['-'.join(i) for i in limit_record[3]]) if 3 in limit_record else '',
                *find_gte_3(limit_record)
                ]
        ws.append(_red_green_blue(data, ws, append_at))
    wb.save(filename=fn)


if __name__ == '__main__':
    # write('20210413',[2700,876,321,32,75,{4:[('邮储','银行'),('阳光城','地产')],3:[('招商','银行'),('金融街','地产'),('獐子岛','养鱼')],5:[('牧原','猪')],1:[('sb','ssb')]}])
    write('20210413')